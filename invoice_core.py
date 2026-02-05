from __future__ import annotations

import re
import math
import os
import zipfile
from dataclasses import dataclass
from datetime import datetime
from collections import Counter
from typing import Iterable, Optional

import pandas as pd
import openpyxl


@dataclass
class GenerateConfig:
    rate_per_hour: float = 100.0
    invoice_date: Optional[datetime] = None  # None => now
    max_line_items: int = 12  # rows 10..21 inclusive in your template
    # Timesheet1.xls has headers on Excel row 2 -> 0-based index 1
    header_row_index: int = 1
    data_row_start_index: int = 2


def _norm_base(name: str) -> str:
    s = str(name).strip().lower()
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # strip "sect"/"section" and anything after
    s = re.split(r"\bsect(?:ion)?\b", s)[0].strip()
    # remove trailing numeric modifiers (e.g., "7 + 9", "1&2", "7-9")
    s = re.sub(r"\s*\b\d+(?:\s*(?:\+|&|and|-)\s*\d+)*\b\s*$", "", s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _choose_display_name(names: Iterable[str]) -> str:
    def score(n: str):
        n = str(n).strip()
        return (len(n.split()), len(n), n.lower())
    return min(names, key=score)


def _build_grouping(project_names: pd.Series) -> tuple[pd.Series, dict[str, str]]:
    base = project_names.apply(_norm_base)
    unique_bases = sorted(set(base.dropna().tolist()), key=lambda x: (len(x.split()), len(x)))

    # group by substring: choose shortest base that is a substring of the longer one
    rep_for: dict[str, str] = {}
    for b in unique_bases:
        candidates = [c for c in unique_bases if c and (c == b or c in b)]
        rep_for[b] = min(candidates, key=lambda x: (len(x.split()), len(x)))

    group_key = base.map(rep_for)

    display_map: dict[str, str] = {}
    tmp = pd.DataFrame({"key": group_key, "name": project_names})
    for key, sub in tmp.groupby("key"):
        display_map[key] = _choose_display_name(sub["name"].dropna().tolist())

    return group_key, display_map


def read_timesheet(timesheet_path: str, cfg: GenerateConfig) -> pd.DataFrame:
    """Reads Timesheet1-style format where row 2 contains the column labels."""
    raw = pd.read_excel(timesheet_path, header=None)
    headers = raw.iloc[cfg.header_row_index].tolist()
    df = raw.iloc[cfg.data_row_start_index:].copy()
    df.columns = headers

    needed = ["Date", "Hours", "Project Name", "Project No"]
    for col in needed:
        if col not in df.columns:
            raise ValueError(f"Missing required column '{col}' in timesheet")

    df = df[needed].copy()
    df = df[df["Date"].notna() & df["Project Name"].notna()].copy()
    df["Date"] = pd.to_datetime(df["Date"])
    df["Hours"] = pd.to_numeric(df["Hours"], errors="coerce").fillna(0)

    return df


def generate_monthly_invoices(
    timesheet_path: str,
    template_path: str,
    output_dir: str,
    cfg: GenerateConfig = GenerateConfig(),
) -> dict:
    """Generates monthly invoices (xlsx) + a zip in output_dir."""
    os.makedirs(output_dir, exist_ok=True)

    df = read_timesheet(timesheet_path, cfg)
    df = df[df["Hours"] > 0].copy()

    df["year"] = df["Date"].dt.year
    df["month"] = df["Date"].dt.month

    group_key, display_map = _build_grouping(df["Project Name"])
    df["group_key"] = group_key

    def is_blank(x):
        return x is None or x == "" or (isinstance(x, float) and math.isnan(x))

    projno_map: dict[str, Optional[int]] = {}
    for key, sub in df.groupby("group_key"):
        vals = [v for v in sub["Project No"].tolist() if not is_blank(v)]
        projno_map[key] = Counter(vals).most_common(1)[0][0] if vals else None

    invoice_date = cfg.invoice_date or datetime.now()

    months = sorted(df[["year", "month"]].drop_duplicates().itertuples(index=False, name=None))
    generated_files: list[str] = []

    for y, m in months:
        sub = df[(df["year"] == y) & (df["month"] == m)].copy()
        if sub.empty:
            continue

        inv_no = int(f"{str(y)[-2:]}{m:02d}")
        first_day = sub["Date"].min().to_pydatetime()
        last_day = sub["Date"].max().to_pydatetime()

        agg = sub.groupby("group_key", as_index=False).agg(Hours=("Hours", "sum"))
        agg["Project Name"] = agg["group_key"].map(display_map)
        agg["Project No"] = agg["group_key"].map(projno_map)
        agg = agg.sort_values(["Project Name"]).reset_index(drop=True)

        agg["Charges"] = agg["Hours"] * float(cfg.rate_per_hour)
        agg["Credits"] = 0.0
        agg["Total Due"] = agg["Charges"]

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Your template locations:
        ws["B5"].value = inv_no           # invoice number at B5
        ws["G1"].value = invoice_date     # invoice date
        ws["G2"].value = first_day
        ws["G3"].value = last_day

        # Clear line items rows 10..21 (B..G)
        start_row = 10
        end_row = start_row + cfg.max_line_items - 1
        for r in range(start_row, end_row + 1):
            for col in ["B", "C", "D", "E", "F", "G"]:
                ws[f"{col}{r}"].value = None

        for i, row in agg.iterrows():
            r = start_row + i
            if r > end_row:
                break
            pn = row["Project No"]
            ws[f"B{r}"].value = None if is_blank(pn) else pn
            ws[f"C{r}"].value = str(row["Project Name"])
            ws[f"D{r}"].value = float(row["Hours"])
            ws[f"E{r}"].value = float(row["Charges"])
            ws[f"F{r}"].value = 0.0
            ws[f"G{r}"].value = float(row["Total Due"])

        out_path = os.path.join(output_dir, f"Invoice_{y}-{m:02d}_{inv_no}.xlsx")
        wb.save(out_path)
        generated_files.append(out_path)

    zip_path = os.path.join(output_dir, "Monthly_Invoices.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for p in generated_files:
            z.write(p, arcname=os.path.basename(p))

    return {"count": len(generated_files), "zip": zip_path, "files": generated_files}
